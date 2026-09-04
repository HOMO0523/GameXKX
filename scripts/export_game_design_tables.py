#!/usr/bin/env python3
"""Export the current GameXXK card and equipment design sources to formatted Excel workbooks."""

from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "design" / "2026-09-04-project-design-tables"
CARD_SOURCE = ROOT / "docs" / "design" / "2026-09-03-all-card-text-review" / "card-texts.json"

QUALITY = [
    (1, "Common", "普通"), (2, "Rare", "稀有"), (3, "Epic", "珍稀"),
    (4, "Legendary", "传奇"), (5, "Immortal", "不朽"), (6, "Treasure", "至宝"),
    (7, "Transcendent", "超凡"), (8, "Celestial", "天界"),
    (9, "Ascendant", "登神"), (10, "Cosmic", "宇宙"),
]
QUALITY_CN = {code: name for _, code, name in QUALITY}
QUALITY_FILL = {
    "普通": "F2F2F2", "稀有": "DDEBF7", "史诗": "E4DFEC", "珍稀": "E4DFEC",
    "传奇": "FFF2CC", "不朽": "FCE4D6", "至宝": "E2F0D9", "超凡": "DDEBF7",
    "天界": "D9EAF7", "登神": "F4CCCC", "宇宙": "D9D2E9",
}

SETS = [
    ("PoJun", "破军"), ("XuanJia", "玄甲"), ("QingNang", "青囊"),
    ("ZhuiFeng", "追风"), ("ShiGu", "蚀骨"), ("ShanHe", "山河"),
]
SLOTS = [
    ("Weapon", "武器"), ("Head", "头冠"), ("Armor", "护甲"),
    ("Belt", "腰带"), ("Shoes", "鞋履"), ("Accessory", "饰品"),
]

CURVES = {
    "Weapon": {"Attack": (2, 1, 1)},
    "Head": {"MaxHealth": (8, 2, 1)},
    "Armor": {"MaxHealth": (4, 1, 1), "Defense": (1, 1, 3)},
    "Belt": {"MaxHealth": (6, 1, 1)},
    "Shoes": {"Speed": (1, 1, 5)},
    "Accessory": {"Attack": (1, 1, 4)},
}

LEGACY = [
    ("Item.IronSword", "青锋短剑", "Weapon", 0, 0, 8, 0, 0, "T_ItemQingfengShortSword"),
    ("Item.ClothArmor", "竹编轻甲", "Armor", 0, 0, 0, 6, 0, "T_ItemBambooLightArmor"),
    ("Item.CranePatternTalisman", "鹤纹护符", "Accessory", 30, 0, 0, 0, 0, "T_ItemCranePatternTalisman"),
    ("Item.InkstonePendant", "墨砚坠饰", "Accessory", 0, 20, 0, 0, 0, "T_ItemInkstonePendant"),
    ("Item.WoodenSword", "木剑", "Weapon", 0, 0, 3, 0, 0, "T_ItemWoodenSword"),
    ("Item.StarterClothArmor", "布甲", "Armor", 0, 0, 0, 3, 0, "T_ItemStarterClothArmor"),
    ("Item.ClothTalisman", "布护符", "Accessory", 10, 0, 0, 0, 0, "T_ItemClothTalisman"),
]

AFFIXES = [
    ("Affix.Universal.MaxHealth", "强身", "通用", "MaxHealth", "BasisPoints", "新装备随机池"),
    ("Affix.Universal.Attack", "劲力", "通用", "Attack", "BasisPoints", "新装备随机池"),
    ("Affix.Universal.Defense", "坚骨", "通用", "Defense", "BasisPoints", "新装备随机池"),
    ("Affix.Universal.MaxMana", "纳息", "历史兼容", "MaxMana", "BasisPoints", "不进入新随机池；当前装备内力结算忽略"),
    ("Affix.Universal.Speed", "轻身", "历史兼容", "Speed", "BasisPoints", "不进入新随机池；旧档可解析并可洗出"),
    ("Affix.PoJun.DirectDamage", "破阵", "破军", "DirectDamage", "BasisPoints", "套装随机池"),
    ("Affix.PoJun.MultiHitDamage", "连锋", "破军", "MultiHitDamage", "BasisPoints", "套装随机池"),
    ("Affix.PoJun.ArmorBreakStacks", "摧甲", "破军", "ArmorBreakStacks", "FlatCount", "套装随机池"),
    ("Affix.PoJun.VulnerableTargetDamage", "乘隙", "破军", "VulnerableTargetDamage", "BasisPoints", "套装随机池"),
    ("Affix.PoJun.FirstAttackDamage", "先声", "破军", "FirstAttackDamage", "BasisPoints", "套装随机池"),
    ("Affix.XuanJia.ArmorGain", "固垒", "玄甲", "ArmorGain", "BasisPoints", "套装随机池"),
    ("Affix.XuanJia.ArmorRetention", "守一", "玄甲", "ArmorRetention", "BasisPoints", "套装随机池"),
    ("Affix.XuanJia.CounterDamage", "反震", "玄甲", "CounterDamage", "BasisPoints", "套装随机池"),
    ("Affix.XuanJia.GuardReduction", "护援", "玄甲", "GuardReduction", "BasisPoints", "套装随机池"),
    ("Affix.XuanJia.LowHealthProtection", "危守", "玄甲", "LowHealthProtection", "BasisPoints", "套装随机池"),
    ("Affix.QingNang.Healing", "回春", "青囊", "Healing", "BasisPoints", "套装随机池"),
    ("Affix.QingNang.Cleanse", "涤尘", "青囊", "Cleanse", "FlatCount", "套装随机池"),
    ("Affix.QingNang.OverhealConversion", "余泽", "青囊", "OverhealConversion", "BasisPoints", "套装随机池"),
    ("Affix.QingNang.ManaRecovery", "养息", "青囊", "ManaRecovery", "BasisPoints", "套装随机池"),
    ("Affix.QingNang.EmergencyHealing", "济危", "青囊", "EmergencyHealing", "BasisPoints", "套装随机池"),
    ("Affix.ZhuiFeng.Draw", "掠影", "追风", "Draw", "FlatCount", "套装随机池"),
    ("Affix.ZhuiFeng.LowCostBonus", "轻策", "追风", "LowCostBonus", "BasisPoints", "套装随机池"),
    ("Affix.ZhuiFeng.SharedEnergy", "聚势", "追风", "SharedEnergy", "FlatCount", "套装随机池"),
    ("Affix.ZhuiFeng.ComboCount", "疾连", "追风", "ComboCount", "FlatCount", "套装随机池"),
    ("Affix.ZhuiFeng.TemporaryCostReduction", "省力", "追风", "TemporaryCostReduction", "FlatCount", "套装随机池"),
    ("Affix.ShiGu.Poison", "淬毒", "蚀骨", "Poison", "FlatCount", "套装随机池"),
    ("Affix.ShiGu.Bleed", "蚀血", "蚀骨", "Bleed", "FlatCount", "套装随机池"),
    ("Affix.ShiGu.Burn", "灼骨", "蚀骨", "Burn", "FlatCount", "套装随机池"),
    ("Affix.ShiGu.DamageOverTime", "绵毒", "蚀骨", "DamageOverTime", "BasisPoints", "套装随机池"),
    ("Affix.ShiGu.StatusRetention", "留煞", "蚀骨", "StatusRetention", "FlatCount", "套装随机池"),
    ("Affix.ShanHe.TerrainPower", "借势", "山河", "TerrainPower", "BasisPoints", "套装随机池"),
    ("Affix.ShanHe.TerrainCostReduction", "循地", "山河", "TerrainCostReduction", "FlatCount", "套装随机池"),
    ("Affix.ShanHe.AdjacentAllyPower", "连营", "山河", "AdjacentAllyPower", "BasisPoints", "套装随机池"),
    ("Affix.ShanHe.FormationPower", "布阵", "山河", "FormationPower", "BasisPoints", "套装随机池"),
    ("Affix.ShanHe.TeamTerrainPower", "山河同势", "山河", "TeamTerrainPower", "BasisPoints", "套装随机池"),
]

MODIFIER_CN = {
    "MaxHealth": "最大生命百分比", "MaxMana": "最大内力百分比（当前忽略）", "Attack": "攻击百分比",
    "Defense": "防御百分比", "Speed": "速度百分比（历史兼容）", "DirectDamage": "直接伤害增幅",
    "MultiHitDamage": "多段伤害增幅", "ArmorBreakStacks": "破甲层数", "VulnerableTargetDamage": "对易伤目标伤害",
    "FirstAttackDamage": "每回合首次攻击伤害", "ArmorGain": "护甲获得增幅", "ArmorRetention": "护甲保留",
    "CounterDamage": "反应伤害", "GuardReduction": "护援减伤", "LowHealthProtection": "低生命保护",
    "Healing": "治疗增幅", "Cleanse": "净化数量", "OverhealConversion": "溢疗转化",
    "ManaRecovery": "内力回复增幅", "EmergencyHealing": "紧急治疗增幅", "Draw": "抽牌数量",
    "LowCostBonus": "低费牌收益", "SharedEnergy": "共享气力回复", "ComboCount": "连击计数",
    "TemporaryCostReduction": "临时减费", "Poison": "中毒点数", "Bleed": "流血点数", "Burn": "灼烧点数",
    "DamageOverTime": "持续伤害增幅", "StatusRetention": "状态保留", "TerrainPower": "地势效果增幅",
    "TerrainCostReduction": "地势联动减费", "AdjacentAllyPower": "相邻友方增幅", "FormationPower": "阵型增幅",
    "TeamTerrainPower": "全队地势增幅",
}

SET_BONUSES = [
    ("Set.PoJun.2", "破军", 2, "穿戴者", "冲锋消费", "每回合首次由穿戴者产生的冲锋被消费后，抽1张牌。", "已实装"),
    ("Set.PoJun.4", "破军", 4, "穿戴者", "收招", "穿戴者收招后，将该牌的冲锋保存为下回合藏式。", "已实装"),
    ("Set.PoJun.6", "破军", 6, "穿戴者", "下回合首张主动牌", "同回合消费冲锋并触发收招：下回合首张主动牌重放基础效果。", "已实装"),
    ("Set.XuanJia.2", "玄甲", 2, "穿戴者", "被动", "获得的护甲提高5%。", "仅描述符；玄甲现行目标与阶位待评审"),
    ("Set.XuanJia.4", "玄甲", 4, "穿戴者", "回合开始", "回合开始保留护甲并使首次直接受击反击80%。", "仅描述符；已确认意图为保留50%护甲，阶位/取整待评审"),
    ("Set.XuanJia.6", "玄甲", 6, "全队", "首次友方生命受伤", "每回合首次有友方受到气血伤害时，为全队提供1次护甲与护援。", "仅描述符；无战斗消费者"),
    ("Set.QingNang.2", "青囊", 2, "全队唯一", "首次2费及以上主动牌", "每回合首次打出2费及以上牌：抽1张牌。", "已实装"),
    ("Set.QingNang.4", "青囊", 4, "全队唯一", "首次2费及以上主动牌", "抽1张；全队非致死失去至多1点生命，再回复2点。", "已实装"),
    ("Set.QingNang.6", "青囊", 6, "全队唯一", "首次2费及以上主动牌", "继承前两档并回复1点共享气力。", "已实装"),
    ("Set.ZhuiFeng.2", "追风", 2, "全队唯一", "主动出牌计数", "全队每主动打出2张牌，抽1张牌。", "已实装"),
    ("Set.ZhuiFeng.4", "追风", 4, "全队唯一", "本回合第2张主动牌", "继承2件；每回合第2张主动牌回复1点气力。", "已实装"),
    ("Set.ZhuiFeng.6", "追风", 6, "全队唯一", "本回合第4张主动牌", "继承前两档；第4张再回1气、全体蓄力1并抽1张。", "已实装"),
    ("Set.ShiGu.2", "蚀骨", 2, "穿戴者", "每卡每目标首次施加DoT", "施加流血、中毒或灼烧时，使该目标获得1点蚀伤。", "已实装"),
    ("Set.ShiGu.4", "蚀骨", 4, "穿戴者", "每回合首次建立双DoT", "目标同时具有至少2种DoT时，自动毒爆1次。", "已实装"),
    ("Set.ShiGu.6", "蚀骨", 6, "穿戴者", "每回合首次毒爆", "首次毒爆不减少流血、中毒和灼烧数值。", "已实装"),
    ("Set.ShanHe.2", "山河", 2, "穿戴者", "被动", "当前地形效果提高5%。", "仅描述符；待评审"),
    ("Set.ShanHe.4", "山河", 4, "穿戴者", "首张地形联动牌", "费用降低1并强化相邻队友。", "仅描述符；相邻/增益/取整待评审"),
    ("Set.ShanHe.6", "山河", 6, "全队", "被动", "当前地形成为阵眼，向全队提供12%对应增益。", "仅描述符；无战斗消费者"),
]


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def curve_value(slot: str, stat: str, level: int) -> int:
    base, num, div = CURVES.get(slot, {}).get(stat, (0, 0, 0))
    level = min(max(level, 1), 20)
    return base if not num else base + ((level - 1) * num) // div


def modern_stats(slot: str, level: int, enhance: int = 0) -> dict[str, int]:
    stats = {k: curve_value(slot, k, level) for k in ("MaxHealth", "MaxMana", "Attack", "Defense", "Speed")}
    factor = 10000 + 1000 * enhance
    for key in ("MaxHealth", "Attack", "Defense", "Speed"):
        stats[key] = stats[key] * factor // 10000
    stats["MaxHealth"] += 2 * enhance
    for key in ("Attack", "Defense", "Speed"):
        stats[key] += enhance
    stats["MaxMana"] = 0
    return stats


def legacy_stats(slot: str, hp: int, mana: int, attack: int, defense: int, speed: int, enhance: int = 0) -> dict[str, int]:
    out = {"MaxHealth": hp, "MaxMana": 0, "Attack": attack, "Defense": defense, "Speed": speed}
    if slot == "Weapon": out["Attack"] += enhance
    elif slot == "Armor": out["Defense"] += enhance
    elif slot == "Accessory": out["Speed"] += enhance
    return out


def safe_text(value):
    if value is None: return ""
    if isinstance(value, (list, tuple, set)): return "\n".join(safe_text(v) for v in value)
    if isinstance(value, dict): return json.dumps(value, ensure_ascii=False, sort_keys=True)
    text = str(value)
    return "'" + text if text[:1] in {"=", "+", "-", "@"} else text


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", color="FFFFFF", bold=True, size=10)
BODY_FONT = Font(name="微软雅黑", size=10)
THIN = Side(style="thin", color="D9E2F3")


def add_table(wb: Workbook, title: str, headers: list[str], rows: list[list], widths: dict[str, int] | None = None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([safe_text(v) if not isinstance(v, (int, float)) else v for v in row])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
            if cell.column in (5, 6) and str(cell.value) in QUALITY_FILL:
                cell.fill = PatternFill("solid", fgColor=QUALITY_FILL[str(cell.value)])
    for idx, header in enumerate(headers, start=1):
        if widths and header in widths:
            width = widths[header]
        else:
            sample = [str(ws.cell(r, idx).value or "") for r in range(1, min(ws.max_row, 80) + 1)]
            width = min(48, max(10, max((max((len(x) for x in s.splitlines()), default=0) for s in sample), default=10) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    return ws


def new_book(title: str) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.creator = "GameXXK / Codex"
    wb.properties.title = title
    wb.properties.subject = "Current design source export"
    return wb


def parse_cost(cost: str):
    match = re.match(r"(\d+)气／(\d+)内(.*)", cost)
    return (int(match.group(1)), int(match.group(2)), match.group(3).lstrip("；")) if match else ("", "", cost)


def build_card_workbook(path: Path) -> dict[str, int]:
    data = json.loads(CARD_SOURCE.read_text(encoding="utf-8"))
    cards = data["cards"]
    wb = new_book("GameXXK 卡牌设计总表")
    group_counts = Counter(c["group"] for c in cards)
    status_counts = Counter(c["status"] for c in cards)
    intro = [["项目", "内容"],
             ["数据口径", "当前173张有效玩家卡；排除25张退役路线牌"],
             ["品质版本", f"{sum(len(c['variants']) for c in cards)}个合法品质版本"],
             ["分支效果", f"{sum(len(c['branches']) for c in cards)}条阵赏/任务分支"],
             ["当前状态", data["meta"]["status"]],
             ["规则实装进度", "171/173主要规则完成；雷走倍率与后巷脱身对象仍待确认"],
             ["待确认CardId", "\n".join(data["meta"]["pending"])],
             ["Tooltip交互", "默认简述；按住Shift详述；按一下Ctrl开关本卡Pill说明"],
             ["显示示例基准", json.dumps(data["meta"]["example"], ensure_ascii=False)],
             ["分组计数", "；".join(f"{k}{v}张" for k, v in group_counts.items())],
             ["状态计数", "；".join(f"{k}{v}张" for k, v in status_counts.items())],
             ["源文件", str(CARD_SOURCE.relative_to(ROOT))],
             ["源SHA256", sha(CARD_SOURCE)]]
    add_table(wb, "00_说明", intro[0], intro[1:], {"项目": 22, "内容": 110})

    master_rows = []
    variant_rows = []
    branch_rows = []
    pill_rows = []
    for seq, card in enumerate(cards, 1):
        qualities = "／".join(v["quality"] for v in card["variants"])
        costs = "\n".join(f"{v['quality']}：{v['cost']}" for v in card["variants"])
        compact = "\n\n".join(f"【{v['quality']}】{v['compact']}" for v in card["variants"])
        detail = "\n\n".join(f"【{v['quality']}】{v['detail']}" for v in card["variants"])
        master_rows.append([seq, card["group"], card["id"], card["name"], card["base_quality"], qualities,
                            card["target_heading"], costs, compact, detail, "、".join(card["pills"]),
                            "\n".join(card["common_rules"]), card["status"], card["note"],
                            card["source_rule"], card["implementation_note"]])
        for variant in card["variants"]:
            energy, mana, extra = parse_cost(variant["cost"])
            pills = "\n".join(f"{p['name']}：{p['description']}" for p in variant.get("pill_descriptions", []))
            variant_rows.append([len(variant_rows) + 1, card["group"], card["id"], card["name"], card["base_quality"],
                                 variant["quality"], energy, mana, extra, variant["cost"], variant["target_heading"],
                                 variant["compact"], variant["detail"], pills, variant.get("pill_shared_note", ""),
                                 card["status"], card["note"], card["source_rule"]])
            for pill in variant.get("pill_descriptions", []):
                pill_rows.append(["品质版本", card["group"], card["id"], card["name"], variant["quality"], "",
                                  pill["name"], "、".join(pill.get("members", [])), pill["description"],
                                  variant.get("pill_shared_note", "")])
        for branch in card["branches"]:
            branch_rows.append([len(branch_rows) + 1, card["group"], card["id"], card["name"], branch["quality"],
                                branch["branch"], branch["target_heading"], branch["compact"], branch["detail"],
                                "\n".join(f"{p['name']}：{p['description']}" for p in branch.get("pill_descriptions", [])),
                                branch.get("pill_shared_note", "")])
            for pill in branch.get("pill_descriptions", []):
                pill_rows.append(["分支", card["group"], card["id"], card["name"], branch["quality"], branch["branch"],
                                  pill["name"], "、".join(pill.get("members", [])), pill["description"],
                                  branch.get("pill_shared_note", "")])

    add_table(wb, "01_卡牌总表",
              ["序号", "分组", "CardId", "卡名", "基础品质", "合法品质", "对象", "费用", "简述", "详述",
               "Pill", "共享规则", "审阅状态", "备注", "取值依据", "实装边界"], master_rows,
              {"CardId": 42, "卡名": 18, "对象": 20, "费用": 28, "简述": 58, "详述": 68,
               "Pill": 25, "共享规则": 48, "备注": 40, "取值依据": 45, "实装边界": 45})
    add_table(wb, "02_品质版本",
              ["序号", "分组", "CardId", "卡名", "基础品质", "当前品质", "气力", "内力", "附加费用规则", "费用原文",
               "对象", "简述", "详述", "本版本Pill说明", "Pill共享注释", "审阅状态", "卡牌备注", "取值依据"], variant_rows,
              {"CardId": 42, "卡名": 18, "附加费用规则": 30, "费用原文": 30, "对象": 22, "简述": 60,
               "详述": 70, "本版本Pill说明": 58, "Pill共享注释": 40, "卡牌备注": 40, "取值依据": 45})
    add_table(wb, "03_分支效果",
              ["序号", "分组", "CardId", "卡名", "品质", "分支", "对象", "简述", "详述", "Pill说明", "共享注释"], branch_rows,
              {"CardId": 42, "卡名": 18, "简述": 62, "详述": 72, "Pill说明": 60, "共享注释": 40})
    add_table(wb, "04_Pill逐卡",
              ["范围", "分组", "CardId", "卡名", "品质", "分支", "Pill", "合并成员", "短说明", "共享注释"], pill_rows,
              {"CardId": 42, "卡名": 18, "短说明": 65, "共享注释": 44})
    add_table(wb, "05_通用术语", ["名称", "类型", "说明", "示例"],
              [[g["name"], g["kind"], g["text"], g["example"]] for g in data["glossary"]],
              {"名称": 24, "类型": 16, "说明": 85, "示例": 38})
    add_table(wb, "06_状态提示", ["提示名", "显示文本", "使用条件"], data["state_texts"],
              {"提示名": 28, "显示文本": 85, "使用条件": 65})
    add_table(wb, "07_术语替换", ["旧称", "现称", "说明"], data["aliases"],
              {"旧称": 35, "现称": 35, "说明": 80})
    formula_rows = [
        ["品质倍率", "普通100%／稀有120%／史诗140%", "连续品质值向上取整", "稀有基础25治疗在100级为150"],
        ["持续伤害增加", "ceil(基础系数×品质%×(队伍最高等级+25)÷2500)", "等级限制1..135", "6点流血、稀有、100级＝36"],
        ["治疗", "ceil((基础治疗+药效)×品质%×(队伍最高等级+25)÷2500)", "所有治疗系数均为原始值", "25点治疗、稀有、100级＝150"],
        ["内力溢出转甲", "ceil(溢出内力×转换百分比×品质%×(等级+25)÷250000)", "仅明确写有内力转换的效果使用", "内力本身不是护甲百分比"],
        ["费用护甲", "ceil(施放者防御×费用系数%×品质%÷10000)", "0/1/2/3+气对应40/80/140/200", "按牌面气力费用"],
        ["DOT上限", "max(25, 25×ceil(队伍最高等级÷25))", "等级限制1..135", "100级上限100"],
        ["等级差伤害", "ceil(伤害×(100+clamp(来源等级−目标等级,-50,50))÷100)", "目标防御与护甲在后续目标结算", "卡牌Tooltip不展开目标结算"],
    ]
    add_table(wb, "08_公式口径", ["项目", "公式", "边界", "示例/说明"], formula_rows,
              {"项目": 24, "公式": 90, "边界": 55, "示例/说明": 65})
    coverage = [["总卡牌", len(cards)], ["品质版本", len(variant_rows)], ["分支效果", len(branch_rows)],
                ["逐卡Pill行", len(pill_rows)], ["通用术语", len(data["glossary"])], ["状态提示", len(data["state_texts"])]]
    coverage += [[f"分组·{k}", v] for k, v in group_counts.items()]
    coverage += [[f"状态·{k}", v] for k, v in status_counts.items()]
    add_table(wb, "09_覆盖核对", ["项目", "数量"], coverage, {"项目": 38, "数量": 16})
    wb.save(path)
    return {"cards": len(cards), "variants": len(variant_rows), "branches": len(branch_rows), "pills": len(pill_rows)}


def build_equipment_workbook(path: Path) -> dict[str, int]:
    wb = new_book("GameXXK 装备设计总表")
    sources = [
        ROOT / "Source/GameXXK/Private/GameXXKEquipmentCatalog.cpp",
        ROOT / "Source/GameXXK/Private/GameXXKAffixCatalog.cpp",
        ROOT / "Source/GameXXK/Private/GameXXKEquipmentSetCatalog.cpp",
        ROOT / "Source/GameXXK/Private/GameXXKGemRules.cpp",
        ROOT / "Source/GameXXK/Private/GameXXKEquipmentRules.cpp",
        ROOT / "Source/GameXXK/Private/GameXXKEquipmentToolRules.cpp",
        ROOT / "docs/design/2026-08-11-gamexxk-project-plan/05-equipment-sets-and-economy.md",
    ]
    intro = [
        ["数据口径", "当前代码目录：42个现代模板（6套装×6部位+6基础装）与7个历史兼容模板"],
        ["固定内力", "装备不得增加角色内力上限；旧墨砚坠饰20内力快照与纳息词缀可读但当前结算忽略"],
        ["品质范围", "装备/词缀/宝石均支持1..10阶；商店直接掉落仅普通/稀有/珍稀，9合1可继续升阶"],
        ["套装状态", "破军、青囊、追风、蚀骨已实装；玄甲、山河仍含待评审描述符"],
        ["属性顺序", "模板等级曲线→强化百分比与固定保底→词缀→宝石→套装战斗效果"],
        ["等级边界", "装备实例允许1..100级，但模板基础曲线在20级封顶"],
        ["仓库", "装备核心集合容量200；桌面仓库按页面显示，不改变核心容量"],
        ["源文件", "\n".join(str(p.relative_to(ROOT)) for p in sources)],
    ]
    add_table(wb, "00_说明", ["项目", "内容"], intro, {"项目": 22, "内容": 115})

    equipment_rows = []
    idx = 0
    for set_code, set_cn in SETS + [("Starter", "基础")]:
        for slot_code, slot_cn in SLOTS:
            idx += 1
            b1 = modern_stats(slot_code, 1)
            b20 = modern_stats(slot_code, 20)
            p10 = modern_stats(slot_code, 20, 10)
            icon_root = "StarterEquipment" if set_code == "Starter" else "Equipment"
            icon_name = f"starter_{slot_code.lower()}" if set_code == "Starter" else f"{set_code.lower()}_{slot_code.lower()}"
            equipment_rows.append([idx, "现代", f"Equipment.{set_code}.{slot_code}", f"{set_cn}{slot_cn}", set_cn, slot_cn,
                                   "ModernPercentBase", b1["MaxHealth"], 0, b1["Attack"], b1["Defense"], b1["Speed"],
                                   b20["MaxHealth"], b20["Attack"], b20["Defense"], b20["Speed"],
                                   p10["MaxHealth"], p10["Attack"], p10["Defense"], p10["Speed"],
                                   f"/Game/GameXXK/UI/{icon_root}/{icon_name}.{icon_name}"])
    slot_cn_map = dict(SLOTS)
    for item_id, name, slot, hp, mana, attack, defense, speed, icon in LEGACY:
        idx += 1
        base = legacy_stats(slot, hp, mana, attack, defense, speed)
        p10 = legacy_stats(slot, hp, mana, attack, defense, speed, 10)
        equipment_rows.append([idx, "历史兼容", item_id, name, "历史", slot_cn_map[slot], "LegacyFlatPerEnhancement",
                               base["MaxHealth"], mana, base["Attack"], base["Defense"], base["Speed"],
                               base["MaxHealth"], base["Attack"], base["Defense"], base["Speed"],
                               p10["MaxHealth"], p10["Attack"], p10["Defense"], p10["Speed"],
                               f"/Game/GameXXK/UI/Inventory/Textures/{icon}.{icon}"])
    add_table(wb, "01_装备总表",
              ["序号", "类型", "模板ID", "名称", "套装", "部位", "缩放规则", "1级生命", "存档内力快照", "1级攻击", "1级防御", "1级速度",
               "20级生命", "20级攻击", "20级防御", "20级速度", "20级+10生命", "20级+10攻击", "20级+10防御", "20级+10速度", "图标路径"],
              equipment_rows, {"模板ID": 40, "名称": 18, "缩放规则": 28, "图标路径": 72})

    growth_rows = []
    for slot, slot_cn in SLOTS:
        formula = []
        for stat, (base, num, div) in CURVES.get(slot, {}).items():
            formula.append(f"{stat}={base}+floor((min(等级,20)-1)×{num}/{div})")
        l1, l10, l20 = modern_stats(slot, 1), modern_stats(slot, 10), modern_stats(slot, 20)
        growth_rows.append([slot_cn, "；".join(formula) or "无模板属性", l1["MaxHealth"], l1["Attack"], l1["Defense"], l1["Speed"],
                            l10["MaxHealth"], l10["Attack"], l10["Defense"], l10["Speed"],
                            l20["MaxHealth"], l20["Attack"], l20["Defense"], l20["Speed"]])
    add_table(wb, "02_部位成长", ["部位", "基础曲线", "L1生命", "L1攻击", "L1防御", "L1速度", "L10生命", "L10攻击", "L10防御", "L10速度", "L20生命", "L20攻击", "L20防御", "L20速度"],
              growth_rows, {"基础曲线": 85})

    quality_rows = []
    tier_rows = []
    for rank, code, cn in QUALITY:
        if rank == 1: weights = "普通100%"
        elif rank == 2: weights = "普通70%／稀有30%"
        else: weights = f"{QUALITY[rank-3][2]}50%／{QUALITY[rank-2][2]}35%／{cn}15%"
        quality_rows.append([rank, code, cn, min(rank, 5), 1 + max(0, rank - 5), weights, "品质不直接倍率化模板基础属性"])
        bp_min = 100 * (rank + 1) * (rank + 2) // 2
        bp_max = bp_min + 100 * (rank + 1)
        tier_rows.append([rank, cn, bp_min, bp_max, bp_min / 100, bp_max / 100, (rank + 1) // 2, rank])
    add_table(wb, "03_品质词缀孔位", ["阶", "代码", "名称", "词缀条数", "宝石孔位", "词缀阶权重", "基础属性说明"], quality_rows,
              {"词缀阶权重": 45, "基础属性说明": 42})

    affix_rows = [[i + 1, *a[:3], a[3], MODIFIER_CN[a[3]], a[4], "百分比" if a[4] == "BasisPoints" else "整数", a[5]] for i, a in enumerate(AFFIXES)]
    add_table(wb, "04_词缀目录", ["序号", "词缀ID", "名称", "归属", "效果族", "效果说明", "存储单位", "显示单位", "池状态"], affix_rows,
              {"词缀ID": 45, "名称": 16, "效果族": 28, "效果说明": 34, "池状态": 48})
    add_table(wb, "05_词缀数值", ["词缀阶", "名称", "基点最小", "基点最大", "百分比最小", "百分比最大", "平坦最小", "平坦最大"], tier_rows)
    add_table(wb, "06_套装效果", ["描述符ID", "套装", "件数", "作用域", "触发", "目录描述/现行意图", "状态"], SET_BONUSES,
              {"描述符ID": 26, "触发": 34, "目录描述/现行意图": 85, "状态": 48})

    gem_rows = []
    gem_types = [("Attack", "攻击", 1), ("Defense", "防御", 1), ("MaxHealth", "生命", 10)]
    for tcode, tcn, base in gem_types:
        for rank, qcode, qcn in QUALITY:
            item = f"Item.Gem.{tcode}.{qcode}"
            gem_rows.append([len(gem_rows) + 1, item, f"{qcn}{tcn}宝石", tcn, rank, qcn, base * (1 << (rank - 1)), "平坦值",
                             f"/Game/GameXXK/UI/Items/Gems/T_Item_Gem_{tcode}_{qcode}.T_Item_Gem_{tcode}_{qcode}"])
    add_table(wb, "07_宝石", ["序号", "道具ID", "名称", "类型", "阶", "品质", "属性增加", "单位", "图标路径"], gem_rows,
              {"道具ID": 38, "名称": 20, "图标路径": 78})

    operation_rows = [
        ["装备", "1件", "实例等级≤角色等级且部位匹配", "替换同槽装备；事务原子提交"],
        ["强化", "1件", "+0..+10；每级1强化石", "现代：floor(B×(1+10%×L))+生命2L/攻防速L；内力永远0"],
        ["洗炼", "1件1词缀", "每次1重铸砂", "重抽同装备合法池；先预览，选择采用或保留；旧Speed可洗出池"],
        ["分解", "1..9件", "未装备且无重复；高品质/强化等需确认", "每件1砂+1强化石+10金币；工具天赋可增金币"],
        ["装备合成", "9件", "同品质且非宇宙", "随机六套之一、随机部位、品质+1；物品等级由工具选定等级区间决定"],
        ["宝石合成", "9个", "同类型同品质且非宇宙", "生成1个下一品质同类型宝石"],
        ["镶嵌", "1装备+1宝石", "孔位有效", "消耗新宝石；被替换旧宝石返还背包"],
        ["永久商店装备包", "100永久金币", "城镇、仓库有空间", "指定套装；部位随机；普通70%/稀有25%/珍稀5%"],
        ["普通历练箱", "每次一项", "50%装备/50%道具", "装备为随机套装/部位普通；道具为三类普通宝石、强化石或砂"],
        ["高级历练箱", "每次一项", "50%装备/50%道具", "装备为随机套装/部位稀有；道具为三类稀有宝石、3强化石或3砂"],
    ]
    add_table(wb, "08_强化工具经济", ["功能", "输入", "条件/成本", "结果"], operation_rows,
              {"功能": 24, "输入": 22, "条件/成本": 62, "结果": 95})
    tool_rows = []
    cumulative = 0
    for level in range(1, 11):
        next_exp = level * 100 if level < 10 else 0
        low, high = (1, 10) if level == 1 else ((level - 1) * 10, level * 10)
        tool_rows.append([level, cumulative, next_exp, low, high])
        cumulative += next_exp
    add_table(wb, "09_工具等级", ["工具等级", "达到本级累计经验", "升下一级经验", "可造装备等级下限", "上限"], tool_rows)
    formulas = [
        ["现代模板属性", "B(L)=一级值+floor((min(L,20)-1)×成长分子÷分母)", "实例等级可到100，曲线20封顶"],
        ["现代强化", "floor(B×(10000+1000×强化级)÷10000)+固定保底", "固定保底：生命2L，攻击/防御/速度各L；内力0"],
        ["旧装备强化", "武器攻击+L；护甲槽防御+L；饰品速度+L", "其他旧槽无强化属性"],
        ["词缀数", "min(装备品质阶,5)", "同实例词缀ID不重复"],
        ["孔位数", "1+max(0,装备品质阶-5)", "1..5阶1孔，10阶6孔"],
        ["词缀基点区间", "min=100(r+1)(r+2)/2；max=min+100(r+1)", "100基点=1%"],
        ["词缀平坦区间", "min=floor((r+1)/2)；max=r", "整数点数"],
        ["工具品质经验", "9^(品质阶-1)", "分解/强化/洗炼/镶嵌按品质；9合1按9倍输入"],
    ]
    add_table(wb, "10_公式与边界", ["项目", "公式", "说明"], formulas, {"项目": 28, "公式": 88, "说明": 75})
    source_rows = [[str(p.relative_to(ROOT)), sha(p)] for p in sources]
    add_table(wb, "11_来源校验", ["源文件", "SHA256"], source_rows, {"源文件": 80, "SHA256": 72})
    wb.save(path)
    return {"equipment": len(equipment_rows), "affixes": len(affix_rows), "sets": len(SET_BONUSES), "gems": len(gem_rows)}


def validate(path: Path, expected: dict[str, int], kind: str):
    wb = load_workbook(path, read_only=True, data_only=False)
    if kind == "card":
        assert wb["01_卡牌总表"].max_row == expected["cards"] + 1
        assert wb["02_品质版本"].max_row == expected["variants"] + 1
        assert wb["03_分支效果"].max_row == expected["branches"] + 1
        assert wb["04_Pill逐卡"].max_row == expected["pills"] + 1
    else:
        assert wb["01_装备总表"].max_row == expected["equipment"] + 1
        assert wb["04_词缀目录"].max_row == expected["affixes"] + 1
        assert wb["06_套装效果"].max_row == expected["sets"] + 1
        assert wb["07_宝石"].max_row == expected["gems"] + 1
    wb.close()


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    card_path = OUT / "GameXXK_卡牌设计总表_2026-09-04.xlsx"
    equipment_path = OUT / "GameXXK_装备设计总表_2026-09-04.xlsx"
    card_counts = build_card_workbook(card_path)
    equipment_counts = build_equipment_workbook(equipment_path)
    validate(card_path, card_counts, "card")
    validate(equipment_path, equipment_counts, "equipment")
    readme = f"""# GameXXK 当前设计总表与配队分析\n\n- `GameXXK_卡牌设计总表_2026-09-04.xlsx`：{card_counts['cards']}张卡、{card_counts['variants']}个品质版本、{card_counts['branches']}条分支、{card_counts['pills']}条逐卡Pill说明。\n- `GameXXK_装备设计总表_2026-09-04.xlsx`：{equipment_counts['equipment']}个装备模板、{equipment_counts['affixes']}条词缀、{equipment_counts['sets']}条套装描述、{equipment_counts['gems']}种宝石。\n- `GameXXK_怪物与阶段数值设计总表_2026-09-04.xlsx`：21种怪物、78个意图、90条意图效果与3个首领第二阶段。\n- `GameXXK_职业配队与伤害期望分析_2026-09-04.html`：3240场队伍组合模拟与2520场正交模拟的交互分析。\n\n卡牌表保留两项未决：`Profession.Sorcerer.RanLingHuanYuan`倍率与`Npc.JinGui.HouXiangTuoShen`对象语义。装备表将玄甲、山河未完成的战斗消费者显式标为待评审。\n\n重新生成：\n\n- `python scripts/export_game_design_tables.py`\n- `python scripts/export_game_enemy_design_table.py`\n- `python scripts/export_game_analysis_html.py`\n"""
    (OUT / "README.md").write_text(readme, encoding="utf-8", newline="\n")
    print(json.dumps({"card": str(card_path), "equipment": str(equipment_path), "card_counts": card_counts, "equipment_counts": equipment_counts}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
