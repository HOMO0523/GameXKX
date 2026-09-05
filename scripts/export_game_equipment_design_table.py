#!/usr/bin/env python3
"""Export the approved post-rebalance equipment design rather than the legacy runtime curves."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from export_game_design_tables import AFFIXES, EQUIPMENT_GLOSSARY, MODIFIER_CN, SETS, SET_BONUSES, SLOTS, add_table, new_book, sha


ROOT = Path(__file__).resolve().parents[1]
SPEC = ROOT / "docs" / "superpowers" / "specs" / "2026-09-03-card-monster-progression-rebalance-design.md"
SET_SPEC = ROOT / "docs" / "superpowers" / "specs" / "2026-09-04-xuanjia-shanhe-set-design.md"
CATALOG = ROOT / "Source" / "GameXXK" / "Private" / "GameXXKEquipmentCatalog.cpp"
SET_CATALOG = ROOT / "Source" / "GameXXK" / "Private" / "GameXXKEquipmentSetCatalog.cpp"

ROLE_STATS = [
    ("Hero", "主角", 2659, 592, 392), ("Blade", "刀客", 2371, 563, 285),
    ("Guard", "守卫", 2800, 425, 358), ("Healer", "药师", 2238, 423, 286),
    ("Hunter", "弓手", 2232, 562, 272), ("Sorcerer", "法师", 2094, 495, 257),
    ("FormationMaster", "阵师", 2374, 452, 301),
]
GEMS = [
    (1,"Common","普通",1,1,5),(2,"Rare","稀有",2,2,10),(3,"Epic","珍稀",4,4,20),
    (4,"Legendary","传奇",8,8,40),(5,"Immortal","不朽",16,16,80),(6,"Treasure","至宝",20,20,100),
    (7,"Transcendent","超凡",25,25,125),(8,"Celestial","天界",32,32,160),
    (9,"Ascendant","登神",40,40,200),(10,"Cosmic","宇宙",50,50,250),
]


def build_approved_equipment_workbook(path: Path) -> dict[str,int]:
    wb=new_book("GameXXK 新装备设计总表")
    intro=[
        ["权威口径","2026-09-03重平衡规格；规格与旧EquipmentCatalog数值冲突时，以规格为准"],
        ["百级基准","角色100级；六件至宝装备；12颗至宝宝石按4攻击/4防御/4生命"],
        ["宝石总增益","+80攻击、+80防御、+400生命；百级最终属性表已经包含，禁止重复相加"],
        ["装备调整","装备攻击贡献约减半；基础防御和生命贡献也减半；精确逐部位新曲线尚待实现落表"],
        ["内力","任何装备、强化、旧饰品和MaxMana词缀均不增加有效内力"],
        ["高阶宝石","不朽后改为约×1.25并向上取整，不再每阶×2"],
        ["装备等级","100级角色可装备101-135级物品；挂机掉落为上一关等级+1至当前关等级"],
        ["实现状态","装备基础曲线/宝石仍按差异表追踪；六套2／4／6件效果已实装，玄甲/山河由v35引入且当前v36保持同一语义"],
    ]
    add_table(wb,"00_说明",["项目","内容"],intro,{"项目":25,"内容":115})
    add_table(wb,"01_百级最终属性",["角色代码","角色","生命","攻击","防御","装备品质","装备件数","宝石配置","状态"],
              [[c,n,h,a,d,"至宝",6,"至宝4攻/4防/4生命","批准冻结基准"] for c,n,h,a,d in ROLE_STATS],{"宝石配置":28,"状态":25})
    gem_rows=[]
    for rank,code,cn,atk,defense,hp in GEMS:
        for t,name,value in (("Attack","攻击",atk),("Defense","防御",defense),("MaxHealth","生命",hp)):
            gem_rows.append([len(gem_rows)+1,f"Item.Gem.{t}.{code}",f"{cn}{name}宝石",rank,cn,name,value,"平坦值","新设计"])
    add_table(wb,"02_宝石总表",["序号","道具ID","名称","阶","品质","属性","数值","单位","口径"],gem_rows,{"道具ID":38,"名称":20})
    template_rows=[]
    for set_code,set_cn in SETS+[("Starter","基础")]:
        for slot_code,slot_cn in SLOTS:
            template_rows.append([len(template_rows)+1,f"Equipment.{set_code}.{slot_code}",f"{set_cn}{slot_cn}",set_cn,slot_cn,
                                  "现代", "新基础生命/攻击/防御贡献约为旧曲线50%；逐部位整数尚待实现冻结", "有效内力0"])
    legacy=[("Item.IronSword","青锋短剑","武器"),("Item.ClothArmor","竹编轻甲","护甲"),("Item.CranePatternTalisman","鹤纹护符","饰品"),("Item.InkstonePendant","墨砚坠饰","饰品"),("Item.WoodenSword","木剑","武器"),("Item.StarterClothArmor","布甲","护甲"),("Item.ClothTalisman","布护符","饰品")]
    for item,name,slot in legacy:
        template_rows.append([len(template_rows)+1,item,name,"历史兼容",slot,"旧档兼容","不作为新装备掉落/百级基准曲线","旧内力快照不生效"])
    add_table(wb,"03_装备模板",["序号","模板ID","名称","套装","部位","类型","新数值口径","内力口径"],template_rows,{"模板ID":42,"新数值口径":78,"内力口径":35})
    affix_rows=[]
    for idx,(aid,name,owner,kind,unit,pool) in enumerate(AFFIXES,1):
        status="兼容保留，不进入新池" if owner=="历史兼容" else "词缀效果保留；幅度需随新装备预算复核"
        if kind=="MaxMana": status="仅存档兼容；当前和新设计均不生效"
        affix_rows.append([idx,aid,name,owner,kind,MODIFIER_CN[kind],unit,pool,status])
    add_table(wb,"04_词缀目录",["序号","词缀ID","名称","归属","效果族","效果说明","单位","旧池状态","新设计状态"],affix_rows,{"词缀ID":45,"效果说明":34,"新设计状态":58})
    add_table(wb,"05_套装效果",["描述符ID","套装","件数","作用域","触发","效果","状态"],SET_BONUSES,{"效果":85,"状态":50})
    add_table(wb,"05A_套装术语",["术语","玩家说明"],EQUIPMENT_GLOSSARY,{"术语":18,"玩家说明":60})
    add_table(wb,"06_至宝配装基准",["项目","数量","单件/单颗","合计","是否已计入角色表"],
              [["至宝装备",6,"按新减半曲线与词缀预算","已折入最终角色属性","是"],["攻击至宝宝石",4,"+20攻击","+80攻击","是"],["防御至宝宝石",4,"+20防御","+80防御","是"],["生命至宝宝石",4,"+100生命","+400生命","是"],["有效内力",0,"装备不提供","0","是"]])
    bands=[]
    levels=list(range(5,136,5))
    for index,level in enumerate(levels):
        low=1 if index==0 else levels[index-1]+1
        bands.append([index+1,("普通" if index<9 else "困难" if index<18 else "地狱"),f"{(index%9)//3+1}-{index%3+1}",level,low,level])
    add_table(wb,"07_掉落等级带",["全局序号","难度","关卡","敌人等级","掉落下限","掉落上限"],bands)
    differences=[
        ["装备基础贡献","当前代码使用旧完整曲线","新设计生命/攻击/防御贡献约减半","待实现"],
        ["宝石生命","当前代码为同阶攻防的10倍","新设计为5倍","待实现"],
        ["不朽后成长","当前代码继续×2","新设计约×1.25并向上取整","待实现"],
        ["至宝孔位","当前代码按品质公式得到2孔","新设计至宝整套共12孔，即每件2孔","一致"],
        ["装备内力","旧ID/快照仍可读","有效贡献恒为0","基础规则已冻结"],
        ["玄甲/山河套装","旧描述符缺少完整消费者","六档批准语义与触发顺序","已实装；v35引入，当前v36保留"],
        ["百级最终属性","旧运行时投影不同","使用01表批准数值","设计冻结、运行时待迁移"],
    ]
    add_table(wb,"08_实现差异",["项目","当前旧实现","批准新设计","状态"],differences,{"当前旧实现":60,"批准新设计":72,"状态":36})
    add_table(wb,"09_来源校验",["源文件","SHA256"],[[str(p.relative_to(ROOT)),sha(p)] for p in (SPEC,SET_SPEC,CATALOG,SET_CATALOG)],{"源文件":85,"SHA256":72})
    wb.save(path)
    return {"equipment":len(template_rows),"affixes":len(affix_rows),"sets":len(SET_BONUSES),"gems":len(gem_rows)}


def main():
    out=ROOT/"docs"/"design"/"2026-09-04-project-design-tables"/"GameXXK_装备设计总表_2026-09-04.xlsx"
    counts=build_approved_equipment_workbook(out)
    wb=load_workbook(out,read_only=True); assert wb["01_百级最终属性"].max_row==8; assert wb["02_宝石总表"].max_row==31; wb.close()
    print(json.dumps({"path":str(out),"counts":counts},ensure_ascii=False,indent=2))


if __name__=="__main__": main()
