#!/usr/bin/env python3
"""Export the current enemy catalog, intents, phases, and encounter numbers to Excel."""

from __future__ import annotations

import json
import math
import re
from pathlib import Path

from openpyxl import load_workbook

from export_game_design_tables import add_table, new_book, sha


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "design" / "2026-09-04-project-design-tables"
CATALOG = ROOT / "Source" / "GameXXK" / "Private" / "GameXXKEnemyCatalog.cpp"
ENCOUNTER = ROOT / "Source" / "GameXXK" / "Private" / "GameXXKEncounterRules.cpp"

TIER_CN = {"Normal": "普通", "Elite": "精英", "Boss": "首领"}
TARGET_CN = {
    "None": "无目标", "Self": "自身", "LowestHealthParty": "最低生命友方",
    "RandomLivingParty": "随机存活友方", "AllLivingParty": "全体存活友方",
    "AllEnemyAllies": "全体敌方友军", "LowestHealthEnemyAlly": "最低生命敌方友军",
    "MarkedParty": "有标记的友方", "PreyTarget": "猎物目标", "MarkedPartyElseRandom": "标记优先，否则稳定随机",
}
STATUS_CN = {
    "None": "无", "Weak": "虚弱", "Mark": "标记", "Bleed": "流血", "Poison": "中毒",
    "Burn": "灼烧", "Medicine": "药效", "Counter": "反击", "Wealth": "聚财",
    "Prey": "猎物", "Rage": "怒气", "Agility": "灵动", "ArmorBreak": "破甲",
}
EFFECT_CN = {
    "DirectDamage": "直接伤害", "AddArmor": "获得护甲", "Heal": "治疗", "ApplyStatus": "施加状态",
    "ConsumeSharedQi": "消耗共享气力", "ModifyAttack": "修改攻击", "ModifyDefense": "修改防御",
    "ModifySpeed": "修改速度", "RemovePositiveStatus": "移除正面状态", "IncreaseNextCardEnergy": "下一张牌气力+1",
    "SetCounter": "设置反击", "SetCharge": "设置蓄势",
}
PASSIVE_CN = {
    "None": "无",
    "IronfeatherFirstHit": "铁羽：每场首次受到直接攻击时，实际生命伤害减半。",
    "BluehornArmorRetention": "护甲留存：敌方阶段开始时保留上一阶段50%护甲。",
    "MoneyRatWealth": "聚财：围绕聚财层、扒窃、散财治疗与钱潮增伤。",
    "PorcupineCounter": "蓄刺反击：通过蓄刺意图建立反击。",
    "GraymaneMarkedHunt": "标记狩猎：攻击有标记目标时，意图伤害提高20%。",
    "RedtuskRage": "怒气：每次受到有效生命伤害获得1怒气，最高5；怒獠每层追加20点伤害。",
    "BlackBearThickHide": "厚皮：受到直接攻击时，生命伤害只保留85%。",
    "WhiteApeStatusGuard": "状态守势：每个玩家回合首次获得状态时，获得8护甲。",
    "DeerHealCooldown": "回春冷却：回春意图触发后按目录冷却2个敌方阶段。",
    "TigerPredator": "捕食者：猎物锁定、虎扑追踪与阶段二流血吸血。",
}
PHASE_CN = {"None": "无", "MoneyRatMadHoard": "守财癫狂", "BlackBearEnraged": "狂怒", "TigerDread": "百兽震惶"}


def split_top(text: str) -> list[str]:
    parts, start, stack, quote, escape = [], 0, [], False, False
    pairs = {"(": ")", "[": "]", "{": "}"}
    for i, ch in enumerate(text):
        if quote:
            if escape: escape = False
            elif ch == "\\": escape = True
            elif ch == '"': quote = False
            continue
        if ch == '"': quote = True
        elif ch in pairs: stack.append(pairs[ch])
        elif stack and ch == stack[-1]: stack.pop()
        elif ch == "," and not stack:
            parts.append(text[start:i].strip()); start = i + 1
    tail = text[start:].strip()
    if tail: parts.append(tail)
    return parts


def calls(text: str, marker: str) -> list[str]:
    result, offset = [], 0
    while True:
        found = text.find(marker, offset)
        if found < 0: return result
        start = found + len(marker)
        depth, quote, escape = 1, False, False
        i = start
        while i < len(text):
            ch = text[i]
            if quote:
                if escape: escape = False
                elif ch == "\\": escape = True
                elif ch == '"': quote = False
            else:
                if ch == '"': quote = True
                elif ch == "(": depth += 1
                elif ch == ")":
                    depth -= 1
                    if depth == 0: break
            i += 1
        if depth != 0: raise ValueError(f"unbalanced call after {marker}")
        result.append(text[start:i]); offset = i + 1


def enum(value: str) -> str:
    return value.strip().rsplit("::", 1)[-1]


def text_value(value: str) -> str:
    match = re.search(r'TEXT\("([^"]*)"\)', value)
    if not match: raise ValueError(f"missing TEXT value: {value}")
    return match.group(1)


def i(value: str, default=0) -> int:
    value = value.strip()
    return int(value) if value else default


def b(value: str, default=False) -> bool:
    value = value.strip()
    return default if not value else value.lower() == "true"


def parse_effect(expr: str) -> dict:
    name, _, rest = expr.strip().partition("(")
    args = split_top(rest[:-1]) if rest.endswith(")") else []
    effect = {"helper": name, "type": "", "target": "", "flat": 0, "attack_pct": 0, "hits": 1,
              "status": "None", "stacks": 0, "consumed": "None", "max_consumed": 0,
              "per_stack": 0, "per_stack_hp_pct": False, "source_status": "None", "source_flat": 0,
              "persistent": False, "phase2_fallback": False, "clear_target": False}
    if name == "Direct":
        effect.update(type="DirectDamage", attack_pct=i(args[0]), target=enum(args[1]) if len(args)>1 else "MarkedPartyElseRandom",
                      hits=i(args[2]) if len(args)>2 else 1, status=enum(args[3]) if len(args)>3 else "None",
                      stacks=i(args[4]) if len(args)>4 else 0)
    elif name == "DirectWithSourceStatusFlatBonus":
        effect.update(type="DirectDamage", attack_pct=i(args[0]), source_status=enum(args[1]), source_flat=i(args[2]),
                      target=enum(args[3]) if len(args)>3 else "MarkedPartyElseRandom")
    elif name == "Armor":
        effect.update(type="AddArmor", flat=i(args[0]), target=enum(args[1]) if len(args)>1 else "Self")
    elif name in {"Status", "PersistentTargetStatus"}:
        effect.update(type="ApplyStatus", status=enum(args[0]), stacks=i(args[1]), target=enum(args[2]), persistent=name.startswith("Persistent"))
    elif name == "PersistentTargetDirect":
        effect.update(type="DirectDamage", attack_pct=i(args[0]), target=enum(args[1]), phase2_fallback=True, clear_target=True)
    elif name == "HealFromConsumedStatus":
        effect.update(type="Heal", target="Self", consumed=enum(args[0]), max_consumed=i(args[1]), per_stack=i(args[2]), per_stack_hp_pct=True)
    elif name == "AttackModifier":
        effect.update(type="ModifyAttack", flat=i(args[0]), target=enum(args[1]))
    elif name == "SpeedModifier":
        effect.update(type="ModifySpeed", flat=i(args[0]), target=enum(args[1]))
    elif name == "MakeEffect":
        effect.update(type=enum(args[0]), target=enum(args[1]), flat=i(args[2]) if len(args)>2 else 0,
                      attack_pct=i(args[3]) if len(args)>3 else 0, hits=i(args[4]) if len(args)>4 else 1,
                      status=enum(args[5]) if len(args)>5 else "None", stacks=i(args[6]) if len(args)>6 else 0)
    else:
        raise ValueError(f"unsupported effect helper: {name}")
    return effect


def effect_text(effect: dict) -> str:
    target = TARGET_CN.get(effect["target"], effect["target"])
    if effect["type"] == "DirectDamage":
        text = f"对{target}造成{effect['attack_pct']}%攻击"
        if effect["hits"] > 1: text += f"×{effect['hits']}段"
        if effect["source_status"] != "None": text += f"；每点自身{STATUS_CN.get(effect['source_status'],effect['source_status'])}+{effect['source_flat']}固定伤害"
        if effect["status"] != "None": text += f"；附加{effect['stacks']}点{STATUS_CN.get(effect['status'],effect['status'])}"
        if effect["persistent"]: text += "；锁定该目标"
        if effect["phase2_fallback"]: text += "；阶段二目标死亡时改取最低生命目标"
        return text
    if effect["type"] == "AddArmor": return f"{target}获得{effect['flat']}护甲"
    if effect["type"] == "ApplyStatus": return f"对{target}施加{effect['stacks']}点{STATUS_CN.get(effect['status'],effect['status'])}" + ("并锁定目标" if effect["persistent"] else "")
    if effect["type"] == "Heal" and effect["consumed"] != "None":
        return f"消耗至多{effect['max_consumed']}点{STATUS_CN.get(effect['consumed'],effect['consumed'])}；每点回复自身最大生命{effect['per_stack']}%"
    if effect["type"] == "Heal": return f"治疗{target}{effect['flat']}点"
    if effect["type"] == "ConsumeSharedQi": return f"消耗玩家共享气力{effect['flat']}点"
    if effect["type"] == "ModifyAttack": return f"{target}攻击+{effect['flat']}"
    if effect["type"] == "ModifyDefense": return f"{target}防御+{effect['flat']}"
    if effect["type"] == "ModifySpeed": return f"{target}速度+{effect['flat']}（按敌方阶段时序）"
    if effect["type"] == "RemovePositiveStatus": return f"移除{target}{effect['flat']}个正面状态"
    if effect["type"] == "IncreaseNextCardEnergy": return "玩家下一张主动牌气力消耗+1"
    return f"{EFFECT_CN.get(effect['type'],effect['type'])} {effect['flat']}"


def parse_catalog() -> list[dict]:
    source = CATALOG.read_text(encoding="utf-8")
    enemies = []
    for raw in calls(source, "Definitions.Add(MakeEnemy("):
        args = split_top(raw)
        enemy = {
            "id": text_value(args[0]), "name": text_value(args[1]), "chapter": i(args[2]), "tier": enum(args[3]),
            "base_hp": i(args[4]), "hp_per": float(args[5].rstrip("f")), "base_attack": i(args[6]),
            "attack_per": float(args[7].rstrip("f")), "base_defense": i(args[8]), "defense_per": float(args[9].rstrip("f")),
            "speed": i(args[10]), "passive": enum(args[12]) if len(args)>12 else "None",
            "phase": enum(args[13]) if len(args)>13 else "None", "round_status": enum(args[14]) if len(args)>14 else "None",
            "round_stacks": i(args[15]) if len(args)>15 else 0, "p2_round_stacks": i(args[16]) if len(args)>16 else 0,
            "p2_direct": i(args[17],100) if len(args)>17 else 100, "p2_attack": i(args[18],100) if len(args)>18 else 100,
            "p2_defense": i(args[19],100) if len(args)>19 else 100,
            "p2_extra_hits": re.findall(r'TEXT\("([^"]+)"\)', args[20]) if len(args)>20 else [],
            "heal_status": enum(args[21]) if len(args)>21 else "None", "heal_missing_pct": i(args[22]) if len(args)>22 else 0,
            "intents": [],
        }
        for intent_raw in calls(args[11], "MakeIntent("):
            ia = split_top(intent_raw)
            effects = [parse_effect(e) for e in split_top(ia[2].strip()[1:-1])]
            enemy["intents"].append({
                "id": text_value(ia[0]), "name": text_value(ia[1]), "effects": effects,
                "charge": i(ia[3]) if len(ia)>3 else 0, "p2_only": b(ia[4]) if len(ia)>4 else False,
                "below_half": b(ia[5]) if len(ia)>5 else False, "required_status": enum(ia[6]) if len(ia)>6 else "None",
                "cooldown": i(ia[7]) if len(ia)>7 else 0, "p2_direct": i(ia[8],100) if len(ia)>8 else 100,
            })
        enemies.append(enemy)
    return enemies


def round_half(value: float) -> int:
    return math.floor(value + .5) if value >= 0 else math.ceil(value - .5)


def stats(enemy: dict, level: int) -> dict:
    level = min(max(level, 1), 100); step = level - 1
    return {"hp": max(1, enemy["base_hp"] + round_half(enemy["hp_per"] * step)),
            "attack": max(1, enemy["base_attack"] + round_half(enemy["attack_per"] * step)),
            "defense": max(0, enemy["base_defense"] + round_half(enemy["defense_per"] * step)), "speed": max(1, enemy["speed"])}


def combat_level(tier: str, route: int) -> int:
    route = min(max(route, 1), 100)
    if tier == "Elite": return min(route + 1, 20)
    if tier == "Boss": return min(route + 2, 20)
    return route


def scale(chapter: int, node: str) -> tuple[int,int,int]:
    if node == "Battle": return 140,250,100
    if node == "Elite": return 160, (270 if chapter==1 else 170 if chapter==2 else 180), (100 if chapter==1 else 105 if chapter==2 else 110)
    return (120 if chapter==1 else 100 if chapter==2 else 80), (120 if chapter==1 else 100 if chapter==2 else 90), 100


def scaled(value: int, percent: int, minimum: int) -> int:
    return max(minimum, (value * percent + 50)//100)


def build(path: Path) -> dict:
    enemies = parse_catalog()
    assert len(enemies) == 21
    assert sum(len(e["intents"]) for e in enemies) == 78
    wb = new_book("GameXXK 怪物与阶段数值设计总表")
    intro = [
        ["目录规模", "21种：每章4普通、2精英、1首领"],
        ["意图规模", "普通每只3个、精英4个、首领6个，共78个"],
        ["阶段", "三个首领在生命≤50%时永久进入第二阶段"],
        ["属性公式", "基础值+四舍五入(每级成长×(战斗等级−1))；生命/攻击最少1，防御最少0"],
        ["等级代码口径", "普通=min(路线等级,100)；精英=min(路线等级+1,20)；首领=min(路线等级+2,20)"],
        ["标准路线快照", "第一章5级、第二章10级、第三章15级，用于本表路线数值预览"],
        ["节点倍率", "普通140%生命/250%攻击；精英160%生命并按章节调整攻击/防御；首领按章节降低倍率"],
        ["源文件", f"{CATALOG.relative_to(ROOT)}\n{ENCOUNTER.relative_to(ROOT)}"],
    ]
    add_table(wb,"00_说明",["项目","内容"],intro,{"项目":24,"内容":115})

    enemy_rows=[]
    for n,e in enumerate(enemies,1):
        s1,s10,s20=stats(e,1),stats(e,10),stats(e,20)
        enemy_rows.append([n,e["chapter"],TIER_CN[e["tier"]],e["id"],e["name"],e["base_hp"],e["hp_per"],e["base_attack"],e["attack_per"],e["base_defense"],e["defense_per"],e["speed"],
                           s10["hp"],s10["attack"],s10["defense"],s20["hp"],s20["attack"],s20["defense"],
                           e["passive"],PASSIVE_CN[e["passive"]],PHASE_CN[e["phase"]],len(e["intents"]),
                           f"/Game/GameXXK/UI/Codex/RouteEnemies/V1/T_{e['id'].replace('.','_')}.T_{e['id'].replace('.','_')}"])
    add_table(wb,"01_怪物总表",["序号","章节","级别","怪物ID","名称","基础生命","生命/级","基础攻击","攻击/级","基础防御","防御/级","速度","L10生命","L10攻击","L10防御","L20生命","L20攻击","L20防御","被动代码","被动说明","第二阶段","意图数","图鉴路径"],enemy_rows,
              {"怪物ID":34,"名称":18,"被动说明":68,"图鉴路径":78})

    level_rows=[]
    for e in enemies:
        for level in (1,5,10,15,20,100):
            s=stats(e,level); level_rows.append([e["chapter"],TIER_CN[e["tier"]],e["id"],e["name"],level,s["hp"],s["attack"],s["defense"],s["speed"]])
    add_table(wb,"02_等级属性",["章节","级别","怪物ID","名称","战斗等级","生命","攻击","防御","速度"],level_rows,{"怪物ID":34})

    route_level={1:5,2:10,3:15}; eligible={"Normal":["Battle","Elite"],"Elite":["Elite","Boss"],"Boss":["Boss"]}; node_cn={"Battle":"普通节点","Elite":"精英节点","Boss":"首领节点"}
    preview=[]
    for e in enemies:
        route=route_level[e["chapter"]]; lvl=combat_level(e["tier"],route); base=stats(e,lvl)
        for node in eligible[e["tier"]]:
            hp_p,at_p,df_p=scale(e["chapter"],node)
            preview.append([e["chapter"],node_cn[node],TIER_CN[e["tier"]],e["name"],route,lvl,hp_p,at_p,df_p,
                            scaled(base["hp"],hp_p,1),scaled(base["attack"],at_p,1),scaled(base["defense"],df_p,0),base["speed"]])
    add_table(wb,"03_路线标准数值",["章节","节点","怪物级别","怪物","路线等级","战斗等级","生命倍率%","攻击倍率%","防御倍率%","最终生命","最终攻击","最终防御","速度"],preview)

    formation=[]
    for chapter in (1,2,3):
        for node in ("Battle","Elite","Boss"):
            hp,at,df=scale(chapter,node)
            layout="1P/3P各随机1只不重复普通怪" if node=="Battle" else "1P/3P普通怪，2P随机精英" if node=="Elite" else "1P/3P固定两只精英，2P首领"
            formation.append([chapter,node_cn[node],layout,"普通=路线等级；精英=min(路线+1,20)；首领=min(路线+2,20)",hp,at,df])
    add_table(wb,"04_编制与倍率",["章节","节点","三槽编制","等级规则","生命倍率%","攻击倍率%","防御倍率%"],formation,{"三槽编制":55,"等级规则":62})

    phase_rows=[]
    for e in [x for x in enemies if x["phase"]!="None"]:
        route=route_level[e["chapter"]]; lvl=combat_level("Boss",route); base=stats(e,lvl); hp_p,at_p,df_p=scale(e["chapter"],"Boss")
        phase_attack=scaled(base["attack"],at_p,1); phase_def=scaled(base["defense"],df_p,0)
        phase_rows.append([e["name"],e["id"],PHASE_CN[e["phase"]],50,"生命≤50%后永久生效",STATUS_CN.get(e["round_status"],e["round_status"]),e["round_stacks"],e["p2_round_stacks"],
                           e["p2_direct"],e["p2_attack"],e["p2_defense"],"、".join(e["p2_extra_hits"]),
                           STATUS_CN.get(e["heal_status"],e["heal_status"]),e["heal_missing_pct"],phase_attack,phase_attack*e["p2_attack"]//100,phase_def,phase_def*e["p2_defense"]//100])
    add_table(wb,"05_首领阶段",["首领","怪物ID","阶段名","阈值%","进入规则","回合开始状态","一阶段层数","二阶段层数","全局直接伤害%","阶段二攻击%","阶段二防御%","追加1段的意图","伤害目标状态","按缺失生命治疗%","标准一阶段攻击","标准二阶段攻击","标准一阶段防御","标准二阶段防御"],phase_rows,
              {"怪物ID":34,"进入规则":35,"追加1段的意图":32})

    intent_rows=[]; effect_rows=[]
    for e in enemies:
        for order,intent in enumerate(e["intents"],1):
            condition=[]
            if intent["p2_only"]: condition.append("仅第二阶段")
            if intent["below_half"]: condition.append("来源生命低于50%")
            if intent["required_status"]!="None": condition.append(f"目标需有{STATUS_CN.get(intent['required_status'],intent['required_status'])}")
            compact="\n".join(effect_text(x) for x in intent["effects"])
            combined=e["p2_direct"]*intent["p2_direct"]//100
            intent_rows.append([e["chapter"],TIER_CN[e["tier"]],e["name"],order,intent["id"],intent["name"],intent["charge"],intent["cooldown"],"；".join(condition) or "常规可用",intent["p2_direct"],e["p2_direct"],combined,"是" if intent["id"] in e["p2_extra_hits"] else "否",compact])
            for effect_index,effect in enumerate(intent["effects"],1):
                effect_rows.append([e["chapter"],TIER_CN[e["tier"]],e["name"],intent["id"],intent["name"],effect_index,EFFECT_CN.get(effect["type"],effect["type"]),TARGET_CN.get(effect["target"],effect["target"]),effect["flat"],effect["attack_pct"],effect["hits"],STATUS_CN.get(effect["status"],effect["status"]),effect["stacks"],STATUS_CN.get(effect["consumed"],effect["consumed"]),effect["max_consumed"],effect["per_stack"],STATUS_CN.get(effect["source_status"],effect["source_status"]),effect["source_flat"],effect_text(effect)])
    add_table(wb,"06_意图总表",["章节","级别","怪物","顺序","意图ID","意图名","蓄势回合","冷却回合","条件","意图阶段二伤害%","怪物阶段二伤害%","合并阶段二伤害%","阶段二追加段数","效果摘要"],intent_rows,
              {"意图ID":25,"意图名":18,"条件":38,"效果摘要":80})
    add_table(wb,"07_意图效果明细",["章节","级别","怪物","意图ID","意图名","效果序号","效果类型","目标","固定值","攻击倍率%","段数","状态","状态点数","消耗状态","最多消耗","每点值","读取自身状态","每点固定加成","效果说明"],effect_rows,
              {"意图ID":25,"效果说明":78})

    passive_rows=[]
    for e in enemies:
        if e["passive"]!="None": passive_rows.append([e["chapter"],TIER_CN[e["tier"]],e["name"],e["passive"],PASSIVE_CN[e["passive"]]])
    add_table(wb,"08_被动机制",["章节","级别","怪物","被动代码","效果"],passive_rows,{"被动代码":32,"效果":90})

    sources=[[str(CATALOG.relative_to(ROOT)),sha(CATALOG)],[str(ENCOUNTER.relative_to(ROOT)),sha(ENCOUNTER)]]
    add_table(wb,"09_来源校验",["源文件","SHA256"],sources,{"源文件":75,"SHA256":72})
    wb.save(path)
    return {"enemies":len(enemies),"intents":len(intent_rows),"effects":len(effect_rows),"phases":len(phase_rows)}


def main():
    OUT.mkdir(parents=True,exist_ok=True)
    path=OUT/"GameXXK_怪物与阶段数值设计总表_2026-09-04.xlsx"
    counts=build(path)
    wb=load_workbook(path,read_only=True)
    assert wb["01_怪物总表"].max_row==22
    assert wb["06_意图总表"].max_row==79
    assert wb["05_首领阶段"].max_row==4
    wb.close()
    print(json.dumps({"path":str(path),"counts":counts},ensure_ascii=False,indent=2))


if __name__=="__main__": main()
